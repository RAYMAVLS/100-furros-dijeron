#!/usr/bin/env python3
"""Convierte un Excel exportado por Microsoft Forms en content/questions.json."""

from __future__ import annotations

import json
import re
import unicodedata
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SETTINGS_PATH = ROOT / "content" / "import-settings.json"
OUTPUT_PATH = ROOT / "content" / "questions.json"


def normalize(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^a-z0-9ñü\s-]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"^(el|la|los|las|un|una|unos|unas|mi|mis)\s+", "", text)
    return text


def display_fallback(value: object) -> str:
    text = re.sub(r"\s+", " ", str(value or "").strip())
    if not text:
        return ""
    return text[0].upper() + text[1:]


def load_json(path: Path, fallback: dict) -> dict:
    if not path.exists():
        return fallback
    with path.open("r", encoding="utf-8") as file:
        return json.load(file)


def resolve_source(path_value: str) -> Path:
    relative = str(path_value or "").lstrip("/")
    return ROOT / relative


def build_aliases(settings: dict) -> dict[str, dict[str, str]]:
    aliases: dict[str, dict[str, str]] = {}
    for item in settings.get("aliases", []):
        question_key = normalize(item.get("question"))
        canonical = str(item.get("canonical") or "").strip()
        if not question_key or not canonical:
            continue
        aliases.setdefault(question_key, {})[normalize(canonical)] = canonical
        for variant in item.get("variants", []):
            key = normalize(variant)
            if key:
                aliases[question_key][key] = canonical
    return aliases


def main() -> None:
    settings = load_json(SETTINGS_PATH, {})
    source = resolve_source(settings.get("sourceFile", "/uploads/respuestas.xlsx"))
    if not source.exists():
        raise FileNotFoundError(
            f"No existe {source.relative_to(ROOT)}. Sube el Excel desde Pages CMS y revisa sourceFile."
        )

    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("El Excel está vacío.")

    headers = [str(value or "").strip() for value in rows[0]]
    ignored = {normalize(item) for item in settings.get("ignoredColumns", [])}
    alias_map = build_aliases(settings)
    top_answers = max(1, int(settings.get("topAnswers", 8)))

    previous = load_json(OUTPUT_PATH, {"questions": []})
    previous_by_text = {
        normalize(question.get("text")): question
        for question in previous.get("questions", [])
        if question.get("text")
    }

    questions = []
    respondent_count = max(0, len(rows) - 1)

    for column_index, header in enumerate(headers):
        question_key = normalize(header)
        if not header or question_key in ignored:
            continue

        counts: Counter[str] = Counter()
        display_names: dict[str, str] = {}
        aliases_for_question = alias_map.get(question_key, {})

        for row in rows[1:]:
            raw = row[column_index] if column_index < len(row) else None
            key = normalize(raw)
            if not key:
                continue
            canonical = aliases_for_question.get(key)
            if canonical:
                canonical_key = normalize(canonical)
                counts[canonical_key] += 1
                display_names[canonical_key] = canonical
            else:
                counts[key] += 1
                display_names.setdefault(key, display_fallback(raw))

        if not counts:
            continue

        previous_question = previous_by_text.get(question_key, {})
        answers = [
            {"text": display_names[key], "points": count}
            for key, count in counts.most_common(top_answers)
        ]
        questions.append({
            "id": previous_question.get("id") or f"q{len(questions) + 1}",
            "enabled": previous_question.get("enabled", True),
            "text": header,
            "multiplier": int(previous_question.get("multiplier", 1)),
            "answers": answers,
        })

    output = {
        "weekLabel": settings.get("weekLabel", "Semana"),
        "respondentCount": respondent_count,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "questions": questions,
    }

    OUTPUT_PATH.write_text(
        json.dumps(output, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"Generadas {len(questions)} preguntas con {respondent_count} respuestas.")


if __name__ == "__main__":
    main()
